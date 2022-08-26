import pdfplumber as plum
import win32com.client
import os
from datetime import datetime
from openpyxl import Workbook,load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from os import path

USER   = os.path.expanduser('~')
FOLDER = USER + "\\Desktop\\COSMETIQUE\\"


def findTable(row):
    try:
        length = 0
        first = row.split()[0]
        if first.startswith(tuple(str(i) for i in range(10))):  #starts with numbering of items before the item code
            length = len(row.split())
            if length > 8 :                                    
                return 1
            else:
                return 0
        else:           
            return 0

    except IndexError :
        return 0


def findItemDetails(row, word_to_find):
    word = ''
    if word_to_find == 'itemcode':
        word = row.split()[0]
    elif word_to_find == 'desc':
        split = row.split()
        length = len(row.split()) - 5            
        for i in range(4,length,1):  
            word = word + ' ' + split[i]
    elif word_to_find == 'qty':        
        word =  row.split()[0]
    elif word_to_find == 'uom':
        split = row.split()
        for i in range(1,4,1):
            word = word + ' ' + split[i]
    elif word_to_find == 'price':
        split = row.split()
        length = len(row.split()) - 5
        word =  split[length]
    elif word_to_find == 'amount':
        split = row.split()
        length = len(row.split()) - 1
        word =  split[length]

    return word


def read_pdf(file_path):  
    global countPage, currentPage, convSuccess
    countPage   = 0
    currentPage = 0
    convSuccess = 0
    with plum.open(file_path) as pdf:
        countPage = len(pdf.pages)  
        for page in pdf.pages:  
            currentPage = currentPage + 1         
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet 1"

            background = PatternFill("solid", start_color="ffff00")
            ws['A1'] = "Order Entry Date:"
            ws['A2'] = "Sales Order No."
            ws['A3'] = "Customer PO No."
            ws['A4'] = "Requested Delivery Date:"
            ws['A5'] = "Sales Invoice No."
            ws['A6'] = "Sold To"
            ws['A8'] = "MATERIAL  CODE"
            ws['B8'] = "CUSTOMER MATERIAL"
            ws['C8'] = "MATERIAL DESCRIPTION"
            ws['D8'] = "QTY"
            ws['E8'] = "UOM"
            ws['F8'] = "UNIT PRICE"
            ws['G8'] = "AMOUNT"
            ws['H8'] = "ADDITIONAL AND DEDUCTIONS"
            ws['I8'] = "AMOUNT"

            sold_to = ""
            sold_add = ""
            si_no  = ""
            # itemcode = ''
            desc    = ''
            qty     = ''
            uom     = ''
            price   = ''
            amount  = ''
            text = page.extract_text() 
            for row in text.split('\n'):

                if si_no == "":
                    if row.startswith("SALES INVOICE Document No.:"):
                        si_no = row.split()[-1]
                        ws['B5'] = si_no
                        ws['B5'].font = Font(name='Courier New',size= 10, bold= True)
                        # ws['B2'] = si_no
                        # ws['B2'].font = Font(name='Courier New',size= 10, bold= True)

                if sold_to == "":
                    if row.startswith('Sold'):
                        split = row.split()
                        le = len(row.split()) - 2
                        for i in range(2,le,1):
                            sold_to = sold_to + ' ' + split[i]
                    continue #skip loop

                if sold_to != '' and sold_add == '':
                    sold_add = row
                
                word = findTable(row)
                if word == 1:
                    length = len(row.split()) 
                    if length >= 10:
                        new_row  = ws.max_row + 1
                        qty      = findItemDetails(row,'qty') 
                        amount   = findItemDetails(row,'amount')
                        uom      = findItemDetails(row,'uom') 
                        price    = findItemDetails(row,'price')                       
                        desc     = findItemDetails(row,'desc') 
                        # itemcode = findItemDetails(row,'itemcode')   

                        # ws['A'+ str(new_row)] = itemcode
                        ws['C'+ str(new_row)] = desc
                        ws['D'+ str(new_row)] = int(qty)
                        ws['E'+ str(new_row)] = uom
                        ws['F'+ str(new_row)] = float(price.replace(",",""))  
                        ws['G'+ str(new_row)] = float(amount.replace(",",""))

                        ws['A' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        ws['C' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        ws['D' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        ws['E' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        ws['F' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        ws['G' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        
            #sum qty
            sum_qty_row = ws.max_row + 3
            last_row    = ws.max_row
            sum_formula = '= SUM(D9:D' + str(last_row) + ')'
            ws['D' + str(sum_qty_row)] = sum_formula

            for i in range(1,7,1):
                ws['A' + str(i)].fill = background
                ws['A' + str(i)].font = Font(name='Courier New',size= 10, bold= True)

            for v in range(1,10,1):
                column_letter = get_column_letter(v)
                ws[column_letter + "8"].fill = background
                ws[column_letter + "8"].font = Font(name='Courier New',size= 10, bold= True)
                ws[column_letter + "8"].alignment = Alignment(horizontal='center')
                if(column_letter == "A" or column_letter == "B"): 
                    ws.column_dimensions[column_letter].width = 30
                elif(column_letter == "F" or column_letter == "G" or column_letter == "I") :
                    ws.column_dimensions[column_letter].width = 15
                elif(column_letter == "C"):
                    ws.column_dimensions[column_letter].width = 80
                elif(column_letter == "D" or column_letter == "E"):
                    ws.column_dimensions[column_letter].width = 10
                elif(column_letter == "H"):
                    ws.column_dimensions[column_letter].width = 30  
        
            if si_no == "" or sold_to == "":
                return 2
            else:
                now = datetime.now()
                date = now.strftime("%m%d%y")
                time = now.strftime("%H%M%S")
                filename =  si_no + '_c.xlsx'        

                if not path.isdir(FOLDER):
                    os.mkdir(FOLDER)

                save_to_path = FOLDER + filename

                wb.save(save_to_path)
                if path.isfile(save_to_path):
                    convert = convert_to_xlsx(save_to_path) 
                    if convert == 1:
                        convSuccess = convSuccess + 1
                    else:
                        pass

                    if currentPage == countPage and convSuccess != 0:
                        return 1
                    elif currentPage == countPage and convSuccess == 0 :
                        return 0

                else:
                    return 0
        #page    
                       
def convert_to_xlsx(file_path): #need to convert again gamit ang excel nga naa sa pc ky dli basahon sa phpexcel ang converted by openpyxl
    getfilename = Path(file_path).stem
    filename = getfilename.replace("_c","")
    save_to = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)
        return 1
    else:
        return 0
                    
                