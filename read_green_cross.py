import win32com.client
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from os import path

USER   = os.path.expanduser('~')
FOLDER = USER + "\\Desktop\\GREEN CROSS\\"

def read_xlsx(file_path):  
    pageToConvert = 0
    convSuccess   = 0
    currentPage   = 0
    wb = load_workbook(file_path, data_only=True)

    for sh in wb.worksheets:
        po = sh['R3'].value
        if po is not None :
            pageToConvert = pageToConvert + 1

    for sheet in wb.worksheets:
        po = sheet['R3'].value
        
        if po is not None :
            currentPage = currentPage + 1   
            itempos = sheet.max_row + 1
            nwb = Workbook()
            ws = nwb.active
            ws.title = "Sheet 1"
            background = PatternFill("solid", start_color="ffff00")
            ws['A1'] = "Order Entry Date:"
            ws['A2'] = "Sales Order No."
            ws['A3'] = "Customer PO No."
            ws['A4'] = "Requested Delivery Date:"
            ws['A5'] = "Sales Invoice No."
            ws['A6'] = "Sold To"
            ws['A8'] = "MATERIAL  CODE"
            ws['B8'] = "CUSTOMER MATERIAL"
            ws['C8'] = "MATERIAL DESCRIPTION"
            ws['D8'] = "QTY"    
            ws['E8'] = "UOM"
            ws['F8'] = "UNIT PRICE"
            ws['G8'] = "AMOUNT"
            ws['H8'] = "ADDITIONAL AND DEDUCTIONS"
            ws['I8'] = "AMOUNT"

            ws['B3'] = po
            ws['B3'].font = Font(name='Courier New',size= 10, bold= True)  

            for i in range(8,itempos, 1):   
                newrow = str(i)
                if sheet['B' + newrow].value is not None and sheet['C' + newrow].value is not None and sheet['D' + newrow].value is not None and sheet['I'+ newrow].value is not None and sheet['R'+ newrow].value != 0 :
                    itemcode = sheet['B' + newrow].value 
                    customer = ""
                    desc     = sheet['C' + newrow].value 
                    qty      = sheet['R' + newrow].value
                    uom      = sheet['D' + newrow].value
                    price    = sheet['I' + newrow].value
                    amount   = sheet['S' + newrow].value

                    new_row = ws.max_row + 1
                    ws['A'+ str(new_row)] = itemcode
                    ws['B'+ str(new_row)] = customer
                    ws['C'+ str(new_row)] = desc
                    ws['D'+ str(new_row)] = qty
                    ws['E'+ str(new_row)] = uom
                    ws['F'+ str(new_row)] = price
                    ws['G'+ str(new_row)] = amount 

                    ws['A' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['B' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['C' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['D' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['E' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['F' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                    ws['G' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)

            #sum qty
            sum_qty_row = ws.max_row + 3
            last_row    = ws.max_row
            sum_formula = '= SUM(D9:D' + str(last_row) + ')'
            ws['D' + str(sum_qty_row)] = sum_formula   

            for i in range(1,7,1):
                ws['A' + str(i)].fill = background
                ws['A' + str(i)].font = Font(name='Courier New',size= 10, bold= True)

            for v in range(1,10,1):
                column_letter = get_column_letter(v)
                ws[column_letter + "8"].fill = background
                ws[column_letter + "8"].font = Font(name='Courier New',size= 10, bold= True)
                ws[column_letter + "8"].alignment = Alignment(horizontal='center')
                if(column_letter == "A" or column_letter == "B"): 
                    ws.column_dimensions[column_letter].width = 30
                elif(column_letter == "F" or column_letter == "G" or column_letter == "I") :
                    ws.column_dimensions[column_letter].width = 15
                elif(column_letter == "C"):
                    ws.column_dimensions[column_letter].width = 80
                elif(column_letter == "D" or column_letter == "E"):
                    ws.column_dimensions[column_letter].width = 10
                elif(column_letter == "H"):
                    ws.column_dimensions[column_letter].width = 30

            now = datetime.now()
            date = now.strftime("%m%d%y")
            time = now.strftime("%H%M%S")
            filename =  po + '_c.xlsx'    
            if not path.isdir(FOLDER):
                os.mkdir(FOLDER)

            save_to_path = FOLDER + filename

            nwb.save(save_to_path)
            if path.isfile(save_to_path):
                convert = convert_to_xlsx_again(save_to_path)  
                if convert == 1:
                    convSuccess = convSuccess + 1
                else:
                    pass

                if currentPage == pageToConvert and convSuccess != 0:
                    return 1
                elif currentPage == pageToConvert and convSuccess == 0 :
                    return 0
            else:
                return 0


def convert_to_xlsx_again(file_path):
    getfilename = Path(file_path).stem
    filename = getfilename.replace("_c","")
    save_to  = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)
        return 1
    else:
        return 0