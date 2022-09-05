import win32com.client
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from os import path

USER     = os.path.expanduser('~')
FOLDER = USER + "\\Desktop\\MEAD JOHNSON\\"

def findItemDetails(row,word_to_find):
    word = ''
    if word_to_find == "itemcode":
        word = row.split()[0]

    elif word_to_find == 'amount':
        word = row.split()[-1]
    
    elif word_to_find == 'price':
        try :
            word    = row.split()[-4]
        except IndexError :
            word = ""
    
    elif word_to_find == 'qty':
        try :
            word      = row.split()[-7]
        except IndexError :
            word = ""
    
    elif word_to_find == 'desc':
        try:
            split = row.split()
            length = len(row.split()) - 9  
            for i in range(1,length,1):  
                word = word + ' ' + split[i]
        except IndexError :
            word = ""

    return word


def read_xlsx(file_path):  
    wb = load_workbook(file_path)
    sheet_obj = wb.active

    si      = ""
    prevItem= ""
    itempos = sheet_obj.max_row + 1

    nwb = Workbook()
    ws = nwb.active
    ws.title = "Sheet 1"
    background = PatternFill("solid", start_color="ffff00")
    ws['A1'] = "Order Entry Date:"
    ws['A2'] = "Sales Order No."
    ws['A3'] = "Customer PO No."
    ws['A4'] = "Requested Delivery Date:"
    ws['A5'] = "Sales Invoice No."
    ws['A6'] = "Sold To"
    ws['A8'] = "MATERIAL  CODE"
    ws['B8'] = "CUSTOMER MATERIAL"
    ws['C8'] = "MATERIAL DESCRIPTION"
    ws['D8'] = "QTY"
    ws['E8'] = "UOM"
    ws['F8'] = "UNIT PRICE"
    ws['G8'] = "AMOUNT"
    ws['H8'] = "ADDITIONAL AND DEDUCTIONS"
    ws['I8'] = "AMOUNT"

    itemcode  = ""
    customer = ""
    desc  = ""
    qty   = ""
    uom   = ""
    price = ""
    amount = ""

    if sheet_obj['A5'].value is not None:
        si = sheet_obj['A5'].value
        ws['B5'] = si
        ws['B5'].font = Font(name='Courier New',size= 10, bold= True)
        ws['B2'] = si
        ws['B2'].font = Font(name='Courier New',size= 10, bold= True)

        for i in range(11,itempos, 1):
           newrow = str(i)
           if sheet_obj['A' + newrow].value is not None :       
                row = sheet_obj['A' + newrow].value
                
                if len(str(row)) > 100 :
                    
                    new_row = ws.max_row + 1
                    itemcode = findItemDetails(row,'itemcode')  
                    qty      = findItemDetails(row,'qty') 
                    uom      = "CS"
                    price    = findItemDetails(row,'price')
                    amount   = findItemDetails(row,'amount')
                    desc     = findItemDetails(row,'desc')  
                    

                    if itemcode != "" and desc != "" and not qty.isalpha() and not price.isalpha() and  not amount.isalpha()  and len(itemcode) >= 5  :

                        if itemcode != prevItem :
                            ws['A'+ str(new_row)] = itemcode
                            ws['C'+ str(new_row)] = desc
                            ws['D'+ str(new_row)] = float(qty)
                            ws['E'+ str(new_row)] = uom
                            ws['F'+ str(new_row)] = float(price.replace(",",""))
                            ws['G'+ str(new_row)] = float(amount.replace(",",""))
                            ws['A' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                            ws['C' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                            ws['D' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                            ws['E' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                            ws['F' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                            ws['G' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                        
                        else :
                            prevRow = new_row - 1
                            prevQty = ws['D'+ str(prevRow)].value
                            prevAmt = ws['G'+ str(prevRow)].value 
                            ws['D'+ str(prevRow)] = prevQty + float(qty)
                            ws['G'+ str(prevRow)] = float(prevAmt) + float(amount.replace(",",""))
                        
                        prevItem = itemcode

        sum_qty_row = ws.max_row + 3
        last_row    = ws.max_row
        sum_formula = '= SUM(D9:D' + str(last_row) + ')'
        ws['D' + str(sum_qty_row)] = sum_formula


        for i in range(1,7,1):
            ws['A' + str(i)].fill = background
            ws['A' + str(i)].font = Font(name='Courier New',size= 10, bold= True)

        for v in range(1,10,1):
            column_letter = get_column_letter(v)
            ws[column_letter + "8"].fill = background
            ws[column_letter + "8"].font = Font(name='Courier New',size= 10, bold= True)
            ws[column_letter + "8"].alignment = Alignment(horizontal='center')
            if(column_letter == "A" or column_letter == "B"): 
                ws.column_dimensions[column_letter].width = 30
            elif(column_letter == "F" or column_letter == "G" or column_letter == "I") :
                ws.column_dimensions[column_letter].width = 15
            elif(column_letter == "C"):
                ws.column_dimensions[column_letter].width = 80
            elif(column_letter == "D" or column_letter == "E"):
                ws.column_dimensions[column_letter].width = 10
            elif(column_letter == "H"):
                ws.column_dimensions[column_letter].width = 30

        now = datetime.now()
        date = now.strftime("%m%d%y")
        time = now.strftime("%H%M%S")
        filename =  str(si)  + '_c.xlsx'    
        if not path.isdir(FOLDER):
            os.mkdir(FOLDER)

        save_to_path = FOLDER + filename

        nwb.save(save_to_path)
        if path.isfile(save_to_path):
            convert = convert_to_xlsx(save_to_path)  
            if convert == 1:
                return 1
            else:
                return 0
        else:
            return 0

    else :
        return 2


def convert_to_xlsx(file_path):
    getfilename = Path(file_path).stem
    filename =  getfilename.replace("_c","")
    save_to  = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)
        return 1
    else:
        return 0



def textfile_to_xlsx(file_path):
    filename = Path(file_path).stem
    save_to  = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)

        result = read_xlsx(save_to)
        if result == 1:
            os.remove(save_to)
            return result
        else:
            return 0

    else:
        return 0
        # print("0")

