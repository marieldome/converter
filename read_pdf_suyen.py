import pdfplumber as plum
import pikepdf
import win32com.client
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from os import path
from db import Database

USER   = os.path.expanduser('~')
FOLDER = USER + "\\Desktop\\SUYEN\\"


#SUPPLIER ID = 10
db        = Database('converter.db')
itemcodes = db.fetch_items_convert(10)


def decrypt_pdf(file_path):
    filename = "SUYEN_decrypted.pdf" 
    if not path.isdir(FOLDER):
        os.mkdir(FOLDER)
    
    with pikepdf.open(file_path) as pdf: #password="salesdept"

        save_to_path = FOLDER + filename
        pdf.save(save_to_path)
        pdf.close()
        if path.isfile(save_to_path):
            
            return save_to_path
            # read_pdf(save_to_path)
        else:
            return "Could not decrypt file!"

def findTable(row): #look for items that start with:
    try:
        length = 0
        first = row.split()[0]
        if first.startswith(tuple(str(i) for i in range(10))):  #starts with numbering of items before the item code
            length = len(row.split())
            if length > 5 :                                    
                second = row.split()[1]
                if second.startswith(tuple(itemcodes)) :                   
                    return 2
                else:
                    return 0
            else:
                return 0

        else:

            if row.startswith(tuple(itemcodes)) : #starts with item code directly
                return 1
            else: 
                return 0
    except IndexError :
        return 0

def findItemDetails(position,row,word_to_find):
        word = ''
        if position == 1:
            if word_to_find == 'itemcode':
                word = row.split()[0]
            elif word_to_find == 'desc':
                split = row.split()
                length = len(row.split()) - 5            
                for i in range(2,length,1):  
                    word = word + ' ' + split[i]
                if word.startswith(tuple(itemcodes)):
                    word = ""
                
            elif word_to_find == 'qty':
                split = row.split()
                length = len(row.split()) - 5
                word =  split[length]
                if word.startswith(tuple(itemcodes)):
                    word = ""

            elif word_to_find == 'uom':
                split = row.split()
                length = len(row.split()) - 4
                word =  split[length]
                if word.startswith(tuple(itemcodes)):
                    word = ""

            elif word_to_find == 'price':
                split = row.split()
                length = len(row.split()) - 3
                word =  split[length]
                if word.startswith(tuple(itemcodes)):
                    word = ""

            elif word_to_find == 'amount':
                split = row.split()
                length = len(row.split()) - 1
                word =  split[length]
                if word.startswith(tuple(itemcodes)):
                    word = ""

        elif position == 2:
            if word_to_find == 'itemcode':
                word = row.split()[1]
                
                # if word.startswith(tuple(itemcodes)):
                #     word = ""
                
                # print(word)

            elif word_to_find == 'desc':
                split = row.split()
                length = len(row.split()) - 5            
                for i in range(3,length,1):  
                    word = word + ' ' + split[i]
                # if word.startswith(tuple(itemcodes)):
                #     word = ""

            elif word_to_find == 'qty':
                split = row.split()
                length = len(row.split()) - 5
                word =  split[length]
                # if word.startswith(tuple(itemcodes)):
                #     word = ""

            elif word_to_find == 'uom':
                split = row.split()
                length = len(row.split()) - 4
                word =  split[length]
                # if word.startswith(tuple(itemcodes)):
                #     word = ""

            elif word_to_find == 'price':
                split = row.split()
                length = len(row.split()) - 3
                word =  split[length]
                # if word.startswith(tuple(itemcodes)):
                #     word = ""

            elif word_to_find == 'amount':
                split = row.split()
                length = len(row.split()) - 1
                word =  split[length]
                # if word.startswith(tuple(itemcodes)):
                #     word = ""

        return word


def read_pdf(file):

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"

    background = PatternFill("solid", start_color="ffff00")
    ws['A1'] = "Order Entry Date:"
    ws['A2'] = "Sales Order No."
    ws['A3'] = "Customer PO No."
    ws['A4'] = "Requested Delivery Date:"
    ws['A5'] = "Sales Invoice No."
    ws['A6'] = "Sold To"
    ws['A8'] = "MATERIAL  CODE"
    ws['B8'] = "CUSTOMER MATERIAL"
    ws['C8'] = "MATERIAL DESCRIPTION"
    ws['D8'] = "QTY"
    ws['E8'] = "UOM"
    ws['F8'] = "UNIT PRICE"
    ws['G8'] = "AMOUNT"
    ws['H8'] = "ADDITIONAL AND DEDUCTIONS"
    ws['I8'] = "AMOUNT"


    order_date = ""
    delivery_date = ""
    reference_no = ""
    itemcode = ""
    desc     = 0.00 
    qty      = 0.00
    uom      = 0.00
    price    = 0.00
    amount   = 0.00


    with plum.open(file) as pdf:
        info = pdf.metadata     

        if not info :   #empty ang metadata sa PSI sa suyen
            for page in pdf.pages:
                text = page.extract_text()
                # print(text)
                for row in text.split('\n'):
                    if order_date == "":
                        if row.startswith('DATE ORDERED'): #DATE ORDERED
                            order_date = row.split()[2]
                            ws['B1'] = order_date
                            ws['B1'].font = Font(name='Courier New',size= 10, bold= True)
                    if delivery_date == "":
                        if row.startswith('DELIVER ON'): #DELIVERY DATE
                            delivery_date = row.split()[2]
                            ws['B4'] = delivery_date
                            ws['B4'].font = Font(name='Courier New',size= 10, bold= True)
                    if reference_no == "":
                        if row.startswith('CANCELLATION'): #REFERENCE
                            reference_no = row.split()[-1]
                            ws['B3'] = reference_no
                            ws['B3'].font = Font(name='Courier New',size= 10, bold= True)

                    
                    position  = findTable(row)
                    if not position == 0 :
                        length = len(row.split())
                        if length > 3:
                            
                            
                            itemcode = findItemDetails(position,row,'itemcode')   
                            desc     = findItemDetails(position,row,'desc')  
                            qty      = findItemDetails(position,row,'qty') 
                            uom      = findItemDetails(position,row,'uom') 
                            price    = findItemDetails(position,row,'price')
                            amount   = findItemDetails(position,row,'amount')

                            # print(itemcode)

                            if not itemcode == "" and not desc == "" and not qty == "" and not uom == "" and not price == "": 
                                new_row = ws.max_row + 1
                                ws['A'+ str(new_row)] = itemcode
                                ws['C'+ str(new_row)] = desc
                                ws['D'+ str(new_row)] = int(qty)
                                ws['E'+ str(new_row)] = uom
                                ws['F'+ str(new_row)] = float(price.replace(",",""))  
                                ws['G'+ str(new_row)] = float(amount.replace(",",""))

                                ws['A' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                                ws['C' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                                ws['D' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                                ws['E' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                                ws['F' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                                ws['G' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
    
            #sum qty
            sum_qty_row = ws.max_row + 3
            last_row    = ws.max_row
            sum_formula = '= SUM(D9:D' + str(last_row) + ')'
            ws['D' + str(sum_qty_row)] = sum_formula

    
            for i in range(1,7,1):
                ws['A' + str(i)].fill = background
                ws['A' + str(i)].font = Font(name='Courier New',size= 10, bold= True)
            
            for v in range(1,10,1):
                column_letter = get_column_letter(v)
                ws[column_letter + "8"].fill = background
                ws[column_letter + "8"].font = Font(name='Courier New',size= 10, bold= True)
                ws[column_letter + "8"].alignment = Alignment(horizontal='center')
                if(column_letter == "A" or column_letter == "B"): 
                    ws.column_dimensions[column_letter].width = 30
                elif(column_letter == "F" or column_letter == "G" or column_letter == "I") :
                    ws.column_dimensions[column_letter].width = 15
                elif(column_letter == "C"):
                    ws.column_dimensions[column_letter].width = 80
                elif(column_letter == "D" or column_letter == "E"):
                    ws.column_dimensions[column_letter].width = 10
                elif(column_letter == "H"):
                    ws.column_dimensions[column_letter].width = 30


            # ws.protection.sheet = True
            # ws.protection.password = "S"+ reference_no + "uyen" #Sheet Password
            # ws.protection.enable()
            # wb.security = WorkbookProtection(workbookPassword = reference_no + "suYEn" , lockStructure = True) #Workbook Password

            
            now = datetime.now()
            date = now.strftime("%m%d%y")
            time = now.strftime("%H%M%S")
            filename = reference_no + '_c.xlsx'    
            if not path.isdir(FOLDER):
                os.mkdir(FOLDER)

            save_to_path = FOLDER + filename

            wb.save(save_to_path)
            if path.isfile(save_to_path):
                convert = convert_to_xlsx(save_to_path)  
                if convert == 1:
                    return 1
                else:
                    return 0
            else:
                return 0

        else :
            return 2


def convert_to_xlsx(file_path):
    getfilename = Path(file_path).stem
    filename = getfilename.replace("_c","")
    save_to  = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)
        return 1
    else:
        return 0