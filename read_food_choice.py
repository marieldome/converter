import win32com.client
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from os import path

USER   = os.path.expanduser('~')
FOLDER = USER + "\\Desktop\\FOOD CHOICE\\"

def read_xlsx(file_path):  
    wb = load_workbook(file_path,data_only=True)
    sheet_obj = wb.active

    soldTo = ""
    sDate  = ""
    itempos    = sheet_obj.max_row + 1
    
    nwb = Workbook()
    ws = nwb.active
    ws.title = "Sheet 1"
    background = PatternFill("solid", start_color="ffff00")
    ws['A1'] = "Order Entry Date:"
    ws['A2'] = "Sales Order No."
    ws['A3'] = "Customer PO No."
    ws['A4'] = "Requested Delivery Date:"
    ws['A5'] = "Sales Invoice No."
    ws['A6'] = "Sold To"
    ws['A8'] = "MATERIAL  CODE"
    ws['B8'] = "CUSTOMER MATERIAL"
    ws['C8'] = "MATERIAL DESCRIPTION"
    ws['D8'] = "QTY"
    ws['E8'] = "UOM"
    ws['F8'] = "UNIT PRICE"
    ws['G8'] = "AMOUNT"
    ws['H8'] = "ADDITIONAL AND DEDUCTIONS"
    ws['I8'] = "AMOUNT"

    if sheet_obj['C2'].value is not None:
        soldTo = sheet_obj['C2'].value
        ws['B6'] = soldTo
        ws['B6'].font = Font(name='Courier New',size= 10, bold= True) 
    if sheet_obj['E2'].value is not None:
        sDate =  sheet_obj['E2'].value
        ws['B1'] = sDate.strftime("%m/%d/%y")
        ws['B1'].font = Font(name='Courier New',size= 10, bold= True)

    if sheet_obj['A8'].value is not None and sheet_obj['B8'].value is not None and sheet_obj['C8'].value is not None and sheet_obj['D8'].value is not None and sheet_obj['E8'].value is not None :
        for i in range(8,itempos, 1):
            newrow = str(i)
            if sheet_obj['A' + newrow].value is not None and sheet_obj['B' + newrow].value is not None and sheet_obj['C' + newrow].value is not None and sheet_obj['D'+ newrow].value is not None and sheet_obj['E' + newrow].value is not None :
                itemcode = ""
                customer = ""
                desc     = sheet_obj['C' + newrow].value 
                qty      = sheet_obj['A' + newrow].value
                uom      = sheet_obj['B' + newrow].value
                price    = sheet_obj['D' + newrow].value
                amount   = sheet_obj['E' + newrow].value
                # if isinstance(price, int) or isinstance(price, float):
                #     amount   = round(qty * price,2)
                # elif price == 'FREE' : 
                #     amount = price 

                new_row = ws.max_row + 1
                ws['A'+ str(new_row)] = itemcode
                ws['B'+ str(new_row)] = customer
                ws['C'+ str(new_row)] = desc
                ws['D'+ str(new_row)] = qty
                ws['E'+ str(new_row)] = uom
                ws['F'+ str(new_row)] = price
                ws['G'+ str(new_row)] = amount     

                ws['A' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['B' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['C' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['D' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['E' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['F' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True)
                ws['G' + str(new_row)].font = Font(name='Courier New',size= 10, bold= True) 

        #sum qty
        sum_qty_row = ws.max_row + 3
        last_row    = ws.max_row
        sum_formula = '= SUM(D9:D' + str(last_row) + ')'
        ws['D' + str(sum_qty_row)] = sum_formula     

        for i in range(1,7,1):
            ws['A' + str(i)].fill = background
            ws['A' + str(i)].font = Font(name='Courier New',size= 10, bold= True)

        for v in range(1,10,1):
            column_letter = get_column_letter(v)
            ws[column_letter + "8"].fill = background
            ws[column_letter + "8"].font = Font(name='Courier New',size= 10, bold= True)
            ws[column_letter + "8"].alignment = Alignment(horizontal='center')
            if(column_letter == "A" or column_letter == "B"): 
                ws.column_dimensions[column_letter].width = 30
            elif(column_letter == "F" or column_letter == "G" or column_letter == "I") :
                ws.column_dimensions[column_letter].width = 15
            elif(column_letter == "C"):
                ws.column_dimensions[column_letter].width = 80
            elif(column_letter == "D" or column_letter == "E"):
                ws.column_dimensions[column_letter].width = 10
            elif(column_letter == "H"):
                ws.column_dimensions[column_letter].width = 30

        now = datetime.now()
        date = now.strftime("%m%d%y")
        time = now.strftime("%H%M%S")
        filename = 'FOODCHOICE_' + '_' + date + time + '_c.xlsx'    
        if not path.isdir(FOLDER):
            os.mkdir(FOLDER)

        save_to_path = FOLDER + filename

        nwb.save(save_to_path)
        if path.isfile(save_to_path):
            convert = convert_to_xlsx_again(save_to_path)  
            if convert == 1:
                return 1
            else:
                return 0
        else:
            return 0


def convert_to_xlsx_again(file_path):
    getfilename = Path(file_path).stem
    filename = getfilename.replace("_c","")
    save_to  = FOLDER + filename + ".xlsx"
    
    excel= win32com.client.Dispatch("Excel.Application") 
    excel.DisplayAlerts = False
    excel.Visible = False
    doc  = excel.Workbooks.Open(file_path)
    doc.SaveAs(save_to, FileFormat = 51)
    excel.DisplayAlerts = True
    excel.Quit()

    if path.isfile(save_to):
        os.remove(file_path)
        return 1
    else:
        return 0